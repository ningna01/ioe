"""
商品相关业务服务
提供商品管理相关的业务逻辑处理
"""
import csv
import io
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Sum

from inventory.models import (
    Product,
    Category,
    ProductImage,
    ProductBatch,
    Warehouse,
    WarehouseInventory,
    UserWarehouseAccess,
)
from inventory.services.warehouse_scope_service import WarehouseScopeService


def _resolve_import_target_warehouse(user):
    if user and getattr(user, 'is_authenticated', False):
        manageable_warehouses = WarehouseScopeService.get_accessible_warehouses(
            user,
            required_permission=UserWarehouseAccess.PERMISSION_PRODUCT_MANAGE,
        )
        preferred_default = WarehouseScopeService.get_default_warehouse(user)
        if preferred_default and manageable_warehouses.filter(id=preferred_default.id).exists():
            return preferred_default
        return manageable_warehouses.first()
    return Warehouse.objects.filter(is_active=True, is_default=True).first() or Warehouse.objects.filter(is_active=True).first()


def _normalize_headers(headers):
    return [str(header or '').strip().lower() for header in headers]


def _build_header_index(headers):
    header_index = {}
    for idx, header in enumerate(headers):
        normalized_header = str(header or '').strip().lower()
        if normalized_header and normalized_header not in header_index:
            header_index[normalized_header] = idx
    return header_index


def _safe_row_value(row, index):
    if index < 0 or index >= len(row):
        return ''
    value = row[index]
    return '' if value is None else str(value).strip()


def _extract_row_value(row, header_index, aliases):
    for alias in aliases:
        idx = header_index.get(alias)
        if idx is None:
            continue
        value = _safe_row_value(row, idx)
        if value:
            return value
    return ''


def _parse_positive_decimal(raw_value):
    normalized = str(raw_value or '').strip().replace(',', '')
    value = Decimal(normalized)
    if value < 0:
        raise InvalidOperation('negative')
    return value


def _parse_non_negative_int(raw_value, *, default=0):
    normalized = str(raw_value or '').strip()
    if not normalized:
        return default
    value = int(normalized)
    if value < 0:
        raise ValueError('negative')
    return value


def _parse_is_active(raw_value, default=True):
    normalized = str(raw_value or '').strip().lower()
    if not normalized:
        return default
    if normalized in {'1', 'true', 'yes', 'y', 'on', '启用', '是', 'active'}:
        return True
    if normalized in {'0', 'false', 'no', 'n', 'off', '禁用', '否', 'inactive'}:
        return False
    return default


def _resolve_category(category_token, default_category):
    normalized_token = str(category_token or '').strip()
    if not normalized_token:
        return default_category

    if normalized_token.isdigit():
        matched = Category.objects.filter(id=int(normalized_token)).first()
        if matched:
            return matched

    category, _ = Category.objects.get_or_create(name=normalized_token)
    return category


def _build_auto_barcode(row_num):
    prefix = timezone.now().strftime('AUTO%Y%m%d%H%M%S')
    candidate = f'{prefix}{row_num:04d}'
    sequence = 1
    while Product.objects.filter(barcode=candidate).exists():
        sequence += 1
        candidate = f'{prefix}{row_num:04d}{sequence:02d}'
    return candidate


def _import_products_from_tabular_data(headers, rows, user):
    result = {
        'strategy': 'row_atomic',
        'success': 0,
        'skipped': 0,
        'failed': 0,
        'failed_rows': []
    }
    target_warehouse = _resolve_import_target_warehouse(user)
    if target_warehouse is None:
        raise ValueError("当前用户没有可用仓库，无法导入商品并初始化库存档案")

    headers_lower = _normalize_headers(headers)
    header_index = _build_header_index(headers_lower)

    missing_headers = []
    if 'name' not in header_index:
        missing_headers.append('name')
    if not any(alias in header_index for alias in ['price', 'retail_price']):
        missing_headers.append('price/retail_price')
    if missing_headers:
        raise ValueError(f"导入文件缺少必要的表头: {', '.join(missing_headers)}")

    default_category, _ = Category.objects.get_or_create(name='未分类')

    for row_num, row in enumerate(rows, start=2):
        try:
            if not row:
                result['skipped'] += 1
                continue

            row_has_values = any(_safe_row_value(row, idx) for idx in range(len(row)))
            if not row_has_values:
                result['skipped'] += 1
                continue

            name = _extract_row_value(row, header_index, ['name'])
            if not name:
                result['failed'] += 1
                result['failed_rows'].append((row_num, "商品名称不能为空"))
                continue

            try:
                retail_price_raw = _extract_row_value(row, header_index, ['price', 'retail_price'])
                retail_price = _parse_positive_decimal(retail_price_raw)
            except (InvalidOperation, ValueError):
                result['failed'] += 1
                result['failed_rows'].append((row_num, "零售价格式不正确"))
                continue

            category = _resolve_category(
                _extract_row_value(row, header_index, ['category', 'category_name', 'category_id']),
                default_category,
            )

            wholesale_price = None
            wholesale_raw = _extract_row_value(row, header_index, ['wholesale_price'])
            if wholesale_raw:
                try:
                    wholesale_candidate = _parse_positive_decimal(wholesale_raw)
                    wholesale_price = wholesale_candidate
                except (InvalidOperation, ValueError):
                    wholesale_price = None

            cost_raw = _extract_row_value(row, header_index, ['cost', 'cost_price'])
            try:
                cost_price = _parse_positive_decimal(cost_raw) if cost_raw else (retail_price * Decimal('0.70'))
            except (InvalidOperation, ValueError):
                cost_price = retail_price * Decimal('0.70')

            barcode = _extract_row_value(row, header_index, ['barcode'])
            if barcode:
                if Product.objects.filter(barcode=barcode).exists():
                    result['skipped'] += 1
                    result['failed_rows'].append((row_num, f"条码 {barcode} 已存在"))
                    continue
            else:
                barcode = _build_auto_barcode(row_num)

            specification = _extract_row_value(row, header_index, ['specification'])
            manufacturer = _extract_row_value(row, header_index, ['manufacturer'])
            description = _extract_row_value(row, header_index, ['description'])
            color = _extract_row_value(row, header_index, ['color'])
            size = _extract_row_value(row, header_index, ['size'])
            is_active = _parse_is_active(
                _extract_row_value(row, header_index, ['is_active', 'active', 'status']),
                default=True,
            )
            try:
                initial_stock = _parse_non_negative_int(
                    _extract_row_value(row, header_index, ['initial_stock', 'initial_quantity', 'opening_stock', 'quantity']),
                    default=0,
                )
            except (TypeError, ValueError):
                result['failed'] += 1
                result['failed_rows'].append((row_num, "初始库存必须是大于等于0的整数"))
                continue

            try:
                warning_level = _parse_non_negative_int(
                    _extract_row_value(row, header_index, ['warning_level', 'warning_stock', 'stock_warning']),
                    default=5,
                )
            except (TypeError, ValueError):
                result['failed'] += 1
                result['failed_rows'].append((row_num, "预警库存必须是大于等于0的整数"))
                continue

            with transaction.atomic():
                product = Product.objects.create(
                    name=name,
                    category=category,
                    price=retail_price,
                    wholesale_price=wholesale_price,
                    cost=cost_price,
                    barcode=barcode,
                    specification=specification,
                    manufacturer=manufacturer,
                    description=description,
                    color=color,
                    size=size,
                    is_active=is_active,
                )

                inventory, _ = WarehouseInventory.objects.get_or_create(
                    product=product,
                    warehouse=target_warehouse,
                    defaults={'warning_level': warning_level, 'quantity': initial_stock}
                )
                inventory_changed_fields = []
                if inventory.warning_level != warning_level:
                    inventory.warning_level = warning_level
                    inventory_changed_fields.append('warning_level')
                if inventory.quantity != initial_stock:
                    inventory.quantity = initial_stock
                    inventory_changed_fields.append('quantity')
                if inventory_changed_fields:
                    inventory.save(update_fields=inventory_changed_fields)

            result['success'] += 1
        except Exception as e:
            result['failed'] += 1
            result['failed_rows'].append((row_num, str(e)))

    return result


def import_products_from_csv(csv_file, user):
    """从CSV文件导入商品"""
    # 读取CSV文件
    raw_content = csv_file.read()
    try:
        decoded_file = raw_content.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            decoded_file = raw_content.decode('gb18030')
        except UnicodeDecodeError as exc:
            raise ValueError(f"CSV 编码错误，请使用 UTF-8 或 GB18030 编码: {exc}") from exc
    csv_data = csv.reader(io.StringIO(decoded_file))
    try:
        headers = next(csv_data)  # 获取表头
    except StopIteration as exc:
        raise ValueError("CSV 文件为空，缺少表头") from exc
    rows = list(csv_data)
    return _import_products_from_tabular_data(headers, rows, user)


def import_products_from_excel(excel_file, user):
    """从 XLSX 文件导入商品。"""
    workbook = load_workbook(excel_file, data_only=True, read_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        headers = next(row_iter)
    except StopIteration as exc:
        raise ValueError("Excel 文件为空，缺少表头") from exc

    rows = [list(row or []) for row in row_iter]
    return _import_products_from_tabular_data(list(headers or []), rows, user)


def search_products(query, category_id=None, active_only=True):
    """搜索商品"""
    products = Product.objects.select_related('category').all()
    
    if query:
        products = products.filter(
            Q(name__icontains=query) | 
            Q(barcode__icontains=query) |
            Q(sku__icontains=query) |
            Q(specification__icontains=query)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if active_only:
        products = products.filter(is_active=True)
    
    return products.order_by('name')


def get_product_with_inventory(product_id):
    """获取商品及其库存信息"""
    try:
        product = Product.objects.get(id=product_id)
        inventory_qs = WarehouseInventory.objects.filter(product=product)
        total_quantity = inventory_qs.aggregate(total=Sum('quantity'))['total'] or 0
        first_inventory = inventory_qs.order_by('warehouse_id').first()
        inventory = {
            'quantity': int(total_quantity),
            'warning_level': first_inventory.warning_level if first_inventory else 10,
        }
        return {
            'product': product,
            'inventory': inventory
        }
    except Product.DoesNotExist:
        return None
