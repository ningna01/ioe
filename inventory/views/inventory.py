"""
库存管理视图
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Sum, F
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from openpyxl import Workbook, load_workbook

import csv
import io

from inventory.models import (
    Product, InventoryTransaction,
    Warehouse, WarehouseInventory,
    OperationLog, StockAlert, check_inventory,
    update_inventory, Category, UserWarehouseAccess
)
from inventory.forms import InventoryTransactionForm
from inventory.services.warehouse_scope_service import WarehouseScopeService


def _ensure_inventory_read_access(user):
    WarehouseScopeService.ensure_any_warehouse_permission(
        user=user,
        required_permission=UserWarehouseAccess.PERMISSION_VIEW,
        error_message='您无权查看库存数据',
    )


def _ensure_inventory_write_access(user, required_permission, error_message):
    WarehouseScopeService.ensure_any_warehouse_permission(
        user=user,
        required_permission=required_permission,
        error_message=error_message,
    )


def _build_inventory_notes(source, intent, user_notes='', extra_context=None):
    """统一库存交易备注格式，便于审计回溯。"""
    note_parts = [
        f"source={source}",
        f"intent={intent}",
    ]
    cleaned_notes = (user_notes or '').strip()
    if cleaned_notes:
        note_parts.append(f"user_note={cleaned_notes}")
    if extra_context:
        for key, value in extra_context.items():
            note_parts.append(f"{key}={value}")
    return " | ".join(note_parts)


def _build_inventory_success_message(action, product, warehouse, delta_quantity, current_quantity):
    return (
        f"{action}成功: {product.name} ({warehouse.name})，"
        f"变更: {delta_quantity:+d}，当前库存: {current_quantity}"
    )


def _build_inventory_failure_message(action, product, warehouse, reason):
    return f"{action}失败: {product.name} ({warehouse.name})，原因: {reason}"


def _get_warehouse_stock(product, warehouse):
    inventory = WarehouseInventory.objects.filter(product=product, warehouse=warehouse).first()
    if inventory is None:
        return 0
    return inventory.quantity


def _build_display_options(raw_values, display_map):
    """将原始值列表转换为下拉可用的 (value, label)。"""
    options = []
    for raw_value in raw_values:
        if raw_value is None:
            continue
        value = str(raw_value).strip()
        if not value:
            continue
        options.append((value, display_map.get(value, value)))
    return options


def _prefill_inventory_form_from_query(request, form):
    """根据 query 参数预填库存操作表单。"""
    product_id = request.GET.get('product_id')
    warehouse_id = request.GET.get('warehouse_id')

    if product_id:
        try:
            pid = int(product_id)
        except (TypeError, ValueError):
            pid = None
        if pid and form.fields['product'].queryset.filter(id=pid).exists():
            form.fields['product'].initial = pid

    if warehouse_id:
        try:
            wid = int(warehouse_id)
        except (TypeError, ValueError):
            wid = None
        if wid and form.fields['warehouse'].queryset.filter(id=wid).exists():
            form.fields['warehouse'].initial = wid


def _normalize_upload_cell(value):
    if value is None:
        return ''
    return str(value).strip()


def _read_tabular_upload(uploaded_file):
    """
    读取 CSV/XLSX 文件并返回规范化数据：
    - headers: 小写列名列表
    - rows: 每行 dict
    """
    file_name = (uploaded_file.name or '').lower()
    if file_name.endswith('.csv'):
        raw_content = uploaded_file.read()
        try:
            decoded_content = raw_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded_content = raw_content.decode('gb18030')
        reader = csv.DictReader(io.StringIO(decoded_content))
        headers = [(_normalize_upload_cell(field)).lower() for field in (reader.fieldnames or [])]
        rows = []
        for raw_row in reader:
            normalized_row = {
                (key or '').strip().lower(): _normalize_upload_cell(value)
                for key, value in (raw_row or {}).items()
            }
            rows.append(normalized_row)
        return headers, rows

    if file_name.endswith('.xlsx'):
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration as exc:
            raise ValueError('上传文件为空') from exc

        headers = [(_normalize_upload_cell(cell)).lower() for cell in (header_row or [])]
        rows = []
        for row_values in row_iter:
            row_dict = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell_value = row_values[index] if row_values and index < len(row_values) else ''
                row_dict[header] = _normalize_upload_cell(cell_value)
            rows.append(row_dict)
        return headers, rows

    raise ValueError('不支持的文件格式，请上传 CSV 或 XLSX 文件')


def _create_inventory_operation_log(
    *,
    operator,
    action,
    product,
    warehouse,
    requested_quantity,
    delta_quantity,
    current_quantity,
    transaction,
    source,
):
    """统一库存操作日志格式。"""
    OperationLog.objects.create(
        operator=operator,
        operation_type='INVENTORY',
        details=(
            f"{action}: 商品={product.name}; 仓库={warehouse.name}; "
            f"请求数量={requested_quantity}; 变更={delta_quantity:+d}; 当前库存={current_quantity}; "
            f"交易ID={transaction.id}; 来源={source}"
        ),
        related_object_id=transaction.id,
        related_content_type=ContentType.objects.get_for_model(InventoryTransaction),
    )


@login_required
def inventory_list(request):
    """库存列表视图（支持按仓库筛选，使用 WarehouseInventory）"""
    _ensure_inventory_read_access(request.user)
    category_id = request.GET.get('category', '')
    color = request.GET.get('color', '')
    size = request.GET.get('size', '')
    search_query = request.GET.get('search', '')
    warehouse_param = request.GET.get('warehouse', '')

    # 仓库筛选：按用户授权解析
    available_warehouses = WarehouseScopeService.get_accessible_warehouses(
        request.user,
        required_permission=UserWarehouseAccess.PERMISSION_VIEW,
    )
    default_warehouse = WarehouseScopeService.get_default_warehouse(request.user)
    if default_warehouse and not WarehouseScopeService.can_access_warehouse(
        request.user,
        default_warehouse,
        required_permission=UserWarehouseAccess.PERMISSION_VIEW,
    ):
        default_warehouse = available_warehouses.first()
    show_all_warehouses = warehouse_param == 'all'
    selected_warehouse = None
    selected_warehouse_value = warehouse_param

    if warehouse_param and not show_all_warehouses:
        try:
            selected_warehouse = available_warehouses.get(id=int(warehouse_param))
        except (ValueError, TypeError):
            selected_warehouse = default_warehouse
            selected_warehouse_value = str(default_warehouse.id) if default_warehouse else ''
        except Warehouse.DoesNotExist:
            selected_warehouse = default_warehouse
            selected_warehouse_value = str(default_warehouse.id) if default_warehouse else ''
    elif show_all_warehouses and not WarehouseScopeService.is_admin_user(request.user):
        # 普通用户仅可查看其授权仓集合
        if not available_warehouses.exists():
            show_all_warehouses = False
            selected_warehouse = None
            selected_warehouse_value = ''
    else:
        selected_warehouse = default_warehouse
        if selected_warehouse is not None:
            selected_warehouse_value = str(selected_warehouse.id)

    # 基础查询：使用 WarehouseInventory（仓库库存唯一真源）
    base_qs = WarehouseInventory.objects.select_related(
        'product', 'product__category', 'warehouse'
    ).all()

    if show_all_warehouses:
        inventory_scope_qs = WarehouseScopeService.filter_warehouse_inventory_queryset(
            request.user,
            base_qs,
            required_permission=UserWarehouseAccess.PERMISSION_VIEW,
        )
    elif selected_warehouse:
        if WarehouseScopeService.can_access_warehouse(
            request.user,
            selected_warehouse,
            required_permission=UserWarehouseAccess.PERMISSION_VIEW,
        ):
            inventory_scope_qs = base_qs.filter(warehouse=selected_warehouse)
        else:
            inventory_scope_qs = base_qs.none()
    else:
        inventory_scope_qs = base_qs.none()

    if category_id:
        inventory_scope_qs = inventory_scope_qs.filter(product__category_id=category_id)
    if search_query:
        inventory_scope_qs = inventory_scope_qs.filter(
            Q(product__name__icontains=search_query) |
            Q(product__barcode__icontains=search_query)
        )

    color_display_map = dict(Product.COLOR_CHOICES)
    size_display_map = dict(Product.SIZE_CHOICES)

    available_color_values = sorted({
        value for value in inventory_scope_qs.values_list('product__color', flat=True)
        if value
    })
    available_size_values = sorted({
        value for value in inventory_scope_qs.values_list('product__size', flat=True)
        if value
    })

    colors = _build_display_options(available_color_values, color_display_map)
    sizes = _build_display_options(available_size_values, size_display_map)

    if color and all(option[0] != color for option in colors):
        colors.append((color, color_display_map.get(color, color)))
    if size and all(option[0] != size for option in sizes):
        sizes.append((size, size_display_map.get(size, size)))

    inventory_items = inventory_scope_qs
    if color:
        inventory_items = inventory_items.filter(product__color=color)
    if size:
        inventory_items = inventory_items.filter(product__size=size)

    categories = Category.objects.all()
    warehouses = available_warehouses

    context = {
        'inventory_items': inventory_items,
        'categories': categories,
        'colors': colors,
        'sizes': sizes,
        'warehouses': warehouses,
        'selected_category': category_id,
        'selected_color': color,
        'selected_size': size,
        'selected_warehouse': selected_warehouse_value,
        'show_all_warehouses': show_all_warehouses,
        'search_query': search_query,
    }
    return render(request, 'inventory/inventory_list.html', context)


@login_required
def inventory_update_warning_level(request, inventory_id):
    """更新仓库库存预警阈值。"""
    if request.method != 'POST':
        messages.error(request, '更新预警库存请通过提交操作完成')
        return redirect('inventory_list')

    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_ADJUST,
        '您无权修改库存预警阈值',
    )

    inventory_item = get_object_or_404(
        WarehouseInventory.objects.select_related('product', 'warehouse'),
        pk=inventory_id,
    )
    WarehouseScopeService.ensure_warehouse_permission(
        user=request.user,
        warehouse=inventory_item.warehouse,
        required_permission=UserWarehouseAccess.PERMISSION_STOCK_ADJUST,
        error_message='您无权修改该仓库的预警库存',
    )

    warning_level_raw = (request.POST.get('warning_level') or '').strip()
    try:
        warning_level = int(warning_level_raw)
    except (TypeError, ValueError):
        warning_level = None

    if warning_level is None or warning_level < 0:
        messages.error(request, '预警库存必须是大于等于 0 的整数')
    else:
        if inventory_item.warning_level != warning_level:
            inventory_item.warning_level = warning_level
            inventory_item.save(update_fields=['warning_level'])

            OperationLog.objects.create(
                operator=request.user,
                operation_type='INVENTORY',
                details=(
                    f"库存预警更新: 商品={inventory_item.product.name}; 仓库={inventory_item.warehouse.name}; "
                    f"新预警={warning_level}; source=inventory_update_warning_level"
                ),
                related_object_id=inventory_item.id,
                related_content_type=ContentType.objects.get_for_model(WarehouseInventory),
            )
        messages.success(
            request,
            f"已更新 {inventory_item.product.name}（{inventory_item.warehouse.name}）预警库存为 {warning_level}"
        )

    next_url = request.POST.get('next', '')
    if isinstance(next_url, str) and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('inventory_list')


@login_required
def inventory_transaction_list(request):
    """库存交易记录列表，显示所有入库、出库和调整记录"""
    _ensure_inventory_read_access(request.user)
    # 获取筛选参数
    transaction_type = request.GET.get('type', '')
    product_id = request.GET.get('product_id', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 基础查询
    transactions = InventoryTransaction.objects.select_related('product', 'operator', 'warehouse').all()
    transactions = WarehouseScopeService.filter_inventory_transactions_queryset(
        request.user,
        transactions,
        required_permission=UserWarehouseAccess.PERMISSION_VIEW,
    )
    
    # 应用筛选条件
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    if product_id:
        transactions = transactions.filter(product_id=product_id)
    
    if search_query:
        transactions = transactions.filter(
            Q(product__name__icontains=search_query) | 
            Q(product__barcode__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    if date_from:
        from datetime import datetime
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            transactions = transactions.filter(created_at__gte=date_from)
        except (ValueError, TypeError):
            pass
    
    if date_to:
        from datetime import datetime, timedelta
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)  # 加一天以包含整天
            transactions = transactions.filter(created_at__lt=date_to)
        except (ValueError, TypeError):
            pass
    
    # 排序
    transactions = transactions.order_by('-created_at')
    
    # 分页
    paginator = Paginator(transactions, 20)  # 每页20条记录
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'inventory/inventory_transaction_list.html', {
        'page_obj': page_obj,
        'transaction_type': transaction_type,
        'product_id': product_id,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'transaction_types': dict(InventoryTransaction.TRANSACTION_TYPES)
    })


@login_required
def inventory_export(request):
    """导出库存快照（CSV / XLSX）。"""
    _ensure_inventory_read_access(request.user)
    export_format = (request.GET.get('format', 'csv') or 'csv').strip().lower()
    selected_warehouse_token = (request.GET.get('warehouse') or '').strip()

    accessible_warehouses = WarehouseScopeService.get_accessible_warehouses(
        request.user,
        required_permission=UserWarehouseAccess.PERMISSION_VIEW,
    )

    inventories = WarehouseInventory.objects.select_related(
        'product', 'product__category', 'warehouse'
    ).filter(warehouse_id__in=accessible_warehouses.values_list('id', flat=True))

    if selected_warehouse_token and selected_warehouse_token != 'all':
        if selected_warehouse_token.isdigit():
            inventories = inventories.filter(warehouse_id=int(selected_warehouse_token))
        else:
            inventories = inventories.filter(warehouse__code=selected_warehouse_token)

    rows = [
        [
            item.warehouse.name,
            item.warehouse.code,
            item.product.name,
            item.product.barcode,
            item.product.category.name if item.product.category else '',
            item.quantity,
            item.warning_level,
            item.product.price,
            item.product.cost,
            item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else '',
        ]
        for item in inventories.order_by('warehouse__name', 'product__name')
    ]
    headers = ['仓库', '仓库编码', '商品名称', '商品条码', '分类', '库存数量', '预警库存', '零售价', '成本价', '更新时间']

    if export_format in ['xlsx', 'excel']:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '库存快照'
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_snapshot.xlsx"'
        return response

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="inventory_snapshot.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


@login_required
def inventory_import(request):
    """批量入库导入（CSV / XLSX）。"""
    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_IN,
        '您无权执行入库操作',
    )

    accessible_warehouses = WarehouseScopeService.get_accessible_warehouses(
        request.user,
        required_permission=UserWarehouseAccess.PERMISSION_STOCK_IN,
    )
    default_warehouse = WarehouseScopeService.get_default_warehouse(request.user)
    if default_warehouse and not accessible_warehouses.filter(id=default_warehouse.id).exists():
        default_warehouse = accessible_warehouses.first()

    if request.method == 'POST':
        upload_file = request.FILES.get('import_file')
        if not upload_file:
            messages.error(request, '请先选择需要导入的 CSV 或 XLSX 文件')
            return render(request, 'inventory/inventory_import.html', {
                'warehouses': accessible_warehouses,
                'default_warehouse_code': default_warehouse.code if default_warehouse else '',
            })

        try:
            headers, rows = _read_tabular_upload(upload_file)
        except Exception as exc:
            messages.error(request, f'文件解析失败: {exc}')
            return render(request, 'inventory/inventory_import.html', {
                'warehouses': accessible_warehouses,
                'default_warehouse_code': default_warehouse.code if default_warehouse else '',
            })

        required_headers = {'barcode', 'quantity'}
        if not required_headers.issubset(set(headers)):
            messages.error(request, '导入文件缺少必要列：barcode, quantity')
            return render(request, 'inventory/inventory_import.html', {
                'warehouses': accessible_warehouses,
                'default_warehouse_code': default_warehouse.code if default_warehouse else '',
            })

        success_count = 0
        failed_count = 0
        failed_rows = []

        for row_index, row in enumerate(rows, start=2):
            barcode = _normalize_upload_cell(row.get('barcode'))
            quantity_raw = _normalize_upload_cell(row.get('quantity'))
            warehouse_code = _normalize_upload_cell(row.get('warehouse_code') or row.get('warehouse'))
            user_note = _normalize_upload_cell(row.get('notes') or row.get('remark'))

            if not barcode:
                failed_count += 1
                failed_rows.append((row_index, 'barcode 不能为空'))
                continue

            try:
                quantity = int(float(quantity_raw))
            except (TypeError, ValueError):
                failed_count += 1
                failed_rows.append((row_index, 'quantity 必须为正整数'))
                continue

            if quantity <= 0:
                failed_count += 1
                failed_rows.append((row_index, 'quantity 必须大于 0'))
                continue

            product = Product.objects.filter(barcode=barcode).first()
            if product is None:
                failed_count += 1
                failed_rows.append((row_index, f'未找到条码为 {barcode} 的商品'))
                continue

            target_warehouse = default_warehouse
            if warehouse_code:
                target_warehouse = accessible_warehouses.filter(code=warehouse_code).first()
                if target_warehouse is None:
                    failed_count += 1
                    failed_rows.append((row_index, f'仓库编码 {warehouse_code} 不存在或无权限'))
                    continue

            if target_warehouse is None:
                failed_count += 1
                failed_rows.append((row_index, '当前用户无可用仓库，请补充 warehouse_code'))
                continue

            notes = _build_inventory_notes(
                source='inventory_import',
                intent='bulk_in',
                user_notes=user_note,
                extra_context={'row': row_index},
            )
            success, inventory_obj, result = update_inventory(
                product=product,
                warehouse=target_warehouse,
                quantity=quantity,
                transaction_type='IN',
                operator=request.user,
                notes=notes,
            )
            if not success:
                failed_count += 1
                failed_rows.append((row_index, str(result)))
                continue

            transaction_obj = result
            _create_inventory_operation_log(
                operator=request.user,
                action='批量入库',
                product=product,
                warehouse=target_warehouse,
                requested_quantity=quantity,
                delta_quantity=quantity,
                current_quantity=inventory_obj.quantity,
                transaction=transaction_obj,
                source='inventory_import',
            )
            success_count += 1

        if success_count:
            messages.success(request, f'批量入库完成：成功 {success_count} 条，失败 {failed_count} 条')
        else:
            messages.warning(request, f'批量入库未写入数据：失败 {failed_count} 条')

        for row_no, reason in failed_rows[:8]:
            messages.warning(request, f'第 {row_no} 行失败：{reason}')
        if len(failed_rows) > 8:
            messages.warning(request, f'其余 {len(failed_rows) - 8} 条失败记录未展开显示')

        return redirect('inventory_list')

    return render(request, 'inventory/inventory_import.html', {
        'warehouses': accessible_warehouses,
        'default_warehouse_code': default_warehouse.code if default_warehouse else '',
    })


@login_required
def inventory_in(request):
    """入库视图（支持多仓库）"""
    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_IN,
        '您无权执行入库操作',
    )
    if request.method == 'POST':
        form = InventoryTransactionForm(
            request.POST,
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_IN,
        )
        if form.is_valid():
            product = form.cleaned_data['product']
            warehouse = form.cleaned_data['warehouse']
            quantity = form.cleaned_data['quantity']
            notes = _build_inventory_notes(
                source='inventory_in',
                intent='manual_in',
                user_notes=form.cleaned_data['notes'],
            )
            success, inventory, result = update_inventory(
                product=product,
                warehouse=warehouse,
                quantity=quantity,
                transaction_type='IN',
                operator=request.user,
                notes=notes
            )
            if success:
                transaction = result
                _create_inventory_operation_log(
                    operator=request.user,
                    action='入库',
                    product=product,
                    warehouse=warehouse,
                    requested_quantity=quantity,
                    delta_quantity=quantity,
                    current_quantity=inventory.quantity,
                    transaction=transaction,
                    source='inventory_in',
                )
                messages.success(
                    request,
                    _build_inventory_success_message(
                        action='入库',
                        product=product,
                        warehouse=warehouse,
                        delta_quantity=quantity,
                        current_quantity=inventory.quantity,
                    ),
                )
                return redirect('inventory_list')
            messages.error(
                request,
                _build_inventory_failure_message(
                    action='入库',
                    product=product,
                    warehouse=warehouse,
                    reason=result,
                ),
            )
    else:
        form = InventoryTransactionForm(
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_IN,
        )
        _prefill_inventory_form_from_query(request, form)
    return render(request, 'inventory/inventory_transaction_form.html', {
        'form': form,
        'form_title': '商品入库',
        'submit_text': '确认入库',
        'transaction_type': 'IN'
    })


@login_required
def inventory_out(request):
    """出库视图（支持多仓库）"""
    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_OUT,
        '您无权执行出库操作',
    )
    if request.method == 'POST':
        form = InventoryTransactionForm(
            request.POST,
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_OUT,
        )
        if form.is_valid():
            product = form.cleaned_data['product']
            warehouse = form.cleaned_data['warehouse']
            quantity = form.cleaned_data['quantity']
            user_notes = form.cleaned_data['notes']
            
            # 先检查库存是否足够（支持多仓库）
            if not check_inventory(product, quantity, warehouse):
                current_quantity = _get_warehouse_stock(product, warehouse)
                messages.error(
                    request,
                    _build_inventory_failure_message(
                        action='出库',
                        product=product,
                        warehouse=warehouse,
                        reason=f'库存不足，当前库存: {current_quantity}，请求出库: {quantity}',
                    ),
                )
                return render(request, 'inventory/inventory_transaction_form.html', {
                    'form': form,
                    'form_title': '商品出库',
                    'submit_text': '确认出库',
                    'transaction_type': 'OUT'
                })

            notes = _build_inventory_notes(
                source='inventory_out',
                intent='manual_out',
                user_notes=user_notes,
            )
            
            # 使用工具函数更新库存
            success, inventory, result = update_inventory(
                product=product,
                warehouse=warehouse,
                quantity=-quantity,  # 负数表示出库
                transaction_type='OUT',
                operator=request.user,
                notes=notes
            )
            
            if success:
                transaction = result
                _create_inventory_operation_log(
                    operator=request.user,
                    action='出库',
                    product=product,
                    warehouse=warehouse,
                    requested_quantity=quantity,
                    delta_quantity=-quantity,
                    current_quantity=inventory.quantity,
                    transaction=transaction,
                    source='inventory_out',
                )
                messages.success(
                    request,
                    _build_inventory_success_message(
                        action='出库',
                        product=product,
                        warehouse=warehouse,
                        delta_quantity=-quantity,
                        current_quantity=inventory.quantity,
                    ),
                )
                return redirect('inventory_list')
            messages.error(
                request,
                _build_inventory_failure_message(
                    action='出库',
                    product=product,
                    warehouse=warehouse,
                    reason=result,
                ),
            )
    else:
        form = InventoryTransactionForm(
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_OUT,
        )
        _prefill_inventory_form_from_query(request, form)
    
    return render(request, 'inventory/inventory_transaction_form.html', {
        'form': form,
        'form_title': '商品出库',
        'submit_text': '确认出库',
        'transaction_type': 'OUT'
    })


@login_required
def inventory_adjust(request):
    """库存调整视图"""
    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_ADJUST,
        '您无权执行库存调整操作',
    )
    if request.method == 'POST':
        form = InventoryTransactionForm(
            request.POST,
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_ADJUST,
        )
        if form.is_valid():
            product = form.cleaned_data['product']
            warehouse = form.cleaned_data['warehouse']
            quantity = form.cleaned_data['quantity']
            notes = form.cleaned_data['notes']
            
            # 获取当前库存
            try:
                inventory = WarehouseInventory.objects.get(product=product, warehouse=warehouse)
                current_quantity = inventory.quantity
            except WarehouseInventory.DoesNotExist:
                current_quantity = 0
            
            # 计算调整值
            adjustment_action = request.POST.get('adjustment_action')
            if adjustment_action == 'set':
                # 设置为指定数量
                if quantity < 0:
                    messages.error(request, '库存数量不能为负数')
                    return render(request, 'inventory/inventory_adjust_form.html', {
                        'form': form,
                        'current_quantity': current_quantity
                    })
                
                adjustment_value = quantity - current_quantity
            elif adjustment_action == 'add':
                # 增加指定数量
                adjustment_value = quantity
            elif adjustment_action == 'subtract':
                # 减少指定数量
                if quantity > current_quantity:
                    messages.error(request, f'减少的数量({quantity})超过了当前库存({current_quantity})')
                    return render(request, 'inventory/inventory_adjust_form.html', {
                        'form': form,
                        'current_quantity': current_quantity
                    })
                
                adjustment_value = -quantity
            else:
                messages.error(request, '请选择有效的调整方式')
                return render(request, 'inventory/inventory_adjust_form.html', {
                    'form': form,
                    'current_quantity': current_quantity
                })
            
            # 使用工具函数更新库存
            notes = _build_inventory_notes(
                source='inventory_adjust',
                intent=f'manual_adjust_{adjustment_action}',
                user_notes=notes,
                extra_context={
                    'before': current_quantity,
                    'delta': f'{adjustment_value:+d}',
                },
            )
            success, inventory, result = update_inventory(
                product=product,
                warehouse=warehouse,
                quantity=adjustment_value,
                transaction_type='ADJUST',
                operator=request.user,
                notes=notes,
            )
            
            if success:
                transaction = result
                _create_inventory_operation_log(
                    operator=request.user,
                    action='调整',
                    product=product,
                    warehouse=warehouse,
                    requested_quantity=quantity,
                    delta_quantity=adjustment_value,
                    current_quantity=inventory.quantity,
                    transaction=transaction,
                    source='inventory_adjust',
                )
                messages.success(
                    request,
                    _build_inventory_success_message(
                        action='调整',
                        product=product,
                        warehouse=warehouse,
                        delta_quantity=adjustment_value,
                        current_quantity=inventory.quantity,
                    ),
                )
                return redirect('inventory_list')
            messages.error(
                request,
                _build_inventory_failure_message(
                    action='调整',
                    product=product,
                    warehouse=warehouse,
                    reason=result,
                ),
            )
    else:
        form = InventoryTransactionForm(
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_ADJUST,
        )
        _prefill_inventory_form_from_query(request, form)
    
    # 获取当前库存（如果已选择商品）
    current_quantity = 0
    selected_product_id = form['product'].value()
    selected_warehouse_id = form['warehouse'].value()
    if selected_product_id and selected_warehouse_id:
        try:
            selected_product_id = int(selected_product_id)
            selected_warehouse_id = int(selected_warehouse_id)
        except (TypeError, ValueError):
            selected_product_id = None
            selected_warehouse_id = None
        if selected_product_id and selected_warehouse_id:
            selected_product = Product.objects.filter(id=selected_product_id).first()
            selected_warehouse = form.fields['warehouse'].queryset.filter(id=selected_warehouse_id).first()
            if selected_product and selected_warehouse:
                try:
                    inventory = WarehouseInventory.objects.get(
                        product=selected_product,
                        warehouse=selected_warehouse,
                    )
                    current_quantity = inventory.quantity
                except WarehouseInventory.DoesNotExist:
                    pass
    
    return render(request, 'inventory/inventory_adjust_form.html', {
        'form': form,
        'current_quantity': current_quantity
    })


@login_required
def inventory_transaction_create(request):
    """创建入库交易视图"""
    _ensure_inventory_write_access(
        request.user,
        UserWarehouseAccess.PERMISSION_STOCK_IN,
        '您无权执行入库操作',
    )
    if request.method == 'POST':
        form = InventoryTransactionForm(
            request.POST,
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_IN,
        )
        if form.is_valid():
            product = form.cleaned_data['product']
            warehouse = form.cleaned_data['warehouse']
            quantity = form.cleaned_data['quantity']
            notes = _build_inventory_notes(
                source='inventory_transaction_create',
                intent='manual_in_legacy_entry',
                user_notes=form.cleaned_data['notes'],
            )

            success, inventory, result = update_inventory(
                product=product,
                warehouse=warehouse,
                quantity=quantity,
                transaction_type='IN',
                operator=request.user,
                notes=notes
            )

            if success:
                transaction = result
                _create_inventory_operation_log(
                    operator=request.user,
                    action='入库',
                    product=product,
                    warehouse=warehouse,
                    requested_quantity=quantity,
                    delta_quantity=quantity,
                    current_quantity=inventory.quantity,
                    transaction=transaction,
                    source='inventory_transaction_create',
                )
                messages.success(
                    request,
                    _build_inventory_success_message(
                        action='入库',
                        product=product,
                        warehouse=warehouse,
                        delta_quantity=quantity,
                        current_quantity=inventory.quantity,
                    ),
                )
                return redirect('inventory_list')

            messages.error(
                request,
                _build_inventory_failure_message(
                    action='入库',
                    product=product,
                    warehouse=warehouse,
                    reason=result,
                ),
            )
    else:
        form = InventoryTransactionForm(
            user=request.user,
            required_permission=UserWarehouseAccess.PERMISSION_STOCK_IN,
        )
    
    return render(request, 'inventory/inventory_form.html', {'form': form}) 
